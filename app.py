import streamlit as st
import pandas as pd
import pdfplumber
import io
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Helper: load CSV with fallback encoding handling
# ---------------------------------------------------------------------------
def load_csv(file_obj):
    """Lê um arquivo CSV tratando possíveis problemas de encoding e separador.
    Tenta UTF-8 e depois Latin-1. Detecta separador ; ou ,
    """
    for enc in ('utf-8', 'latin1'):
        for sep in (';', ','):
            try:
                file_obj.seek(0)
                df = pd.read_csv(file_obj, encoding=enc, sep=sep)
                if len(df.columns) > 1:  # se só 1 coluna, separador errado
                    return df
            except Exception:
                continue
    # último recurso
    file_obj.seek(0)
    return pd.read_csv(file_obj, encoding='latin1', sep=';')


st.set_page_config(page_title="Analisador de Notas Fiscais", layout="wide", page_icon="📄")

def clean_currency(x):
    """Converte valores monetários do formato BRL para float."""
    if isinstance(x, str):
        # Remove R$, pontos de milhar e substitui vírgula por ponto
        x = x.replace('R$', '').replace('.', '').replace(',', '.').strip()
        try:
            return float(x)
        except ValueError:
            return 0.0
    try:
        return float(x)
    except (ValueError, TypeError):
        return 0.0

@st.cache_data
def process_pdf(file_bytes):
    """Extrai tabelas do PDF e retorna um DataFrame consolidado."""
    all_data = []
    headers = None
    
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        total_pages = len(pdf.pages)
        progress_bar = st.progress(0)
        
        for i, page in enumerate(pdf.pages):
            # Mostrando o progresso
            progress_bar.progress((i + 1) / total_pages, text=f"Lendo página {i + 1} de {total_pages}...")
            
            # Extraindo a tabela
            tables = page.extract_tables()
            for table in tables:
                if not table: continue
                
                # Assume-se que a primeira página tem o cabeçalho
                # Para tabelas de múltiplas páginas, precisamos lidar com repetição de cabeçalho
                for row in table:
                    # Limpa quebras de linha nas células
                    clean_row = [str(cell).replace('\n', ' ') if cell else '' for cell in row]
                    
                    # Identifica cabeçalho se ainda não o tivermos
                    if headers is None and 'Nota' in clean_row[0] and 'Tomador' in str(clean_row):
                        headers = clean_row
                        continue
                        
                    # Pula repetição de cabeçalhos nas páginas subsequentes
                    if headers and clean_row == headers:
                        continue
                        
                    # Se for uma linha de dados e já tivermos o cabeçalho
                    if headers and len(clean_row) == len(headers):
                        primeira_coluna = clean_row[0].strip()
                        
                        # Se a primeira coluna tem números (ex: "00003846"), é uma nota nova
                        if any(char.isdigit() for char in primeira_coluna):
                            all_data.append(clean_row)
                        # Se não tem números, mas tem algum texto nas outras colunas, é provável continuação
                        elif all_data and any(c.strip() for c in clean_row):
                            # É continuação da linha anterior, então vamos juntar os textos
                            for j in range(len(headers)):
                                if clean_row[j].strip():
                                    all_data[-1][j] += " " + clean_row[j].strip()

    if headers and all_data:
        df = pd.DataFrame(all_data, columns=headers)
        return df
    else:
        return pd.DataFrame()

def prepare_dataframe(df):
    """Limpa e formata as colunas principais. Suporta CSV e PDF."""

    # --- Detecta se é CSV NFS-e pela presença de coluna específica ---
    all_cols = [c.strip() for c in df.columns]
    is_csv = any('CPF/CNPJ/NIF do Tomador' in c for c in all_cols)

    if is_csv:
        # Mapeamento EXATO para CSV - seleciona apenas as colunas necessárias
        csv_col_map = {
            'da Nota Fiscal Eletr':    'Nota',
            'Data Hora da Emiss':      'Emissão',
            'CPF/CNPJ/NIF do Tomador': 'CNPJ',
            'Social do Tomador':       'Tomador de Serviços',
            'Al\u00edquota':           'Alíquota',
            'Valor dos Servi':         'Valor da Nota',
            'Valor do ISS':            'Valor ISS',
        }
        # Encontra quais colunas originais batem com o mapa
        rename_map = {}
        for orig_col in df.columns:
            clean = orig_col.strip()
            for pattern, canonical in csv_col_map.items():
                if pattern in clean:
                    rename_map[orig_col] = canonical
                    break

        # Mantém apenas as colunas que foram mapeadas
        df = df[list(rename_map.keys())].rename(columns=rename_map)

        # Limpeza de tipos
        if 'Valor da Nota' in df.columns:
            df['Valor da Nota'] = df['Valor da Nota'].apply(clean_currency)
        if 'Valor ISS' in df.columns:
            df['Valor ISS'] = df['Valor ISS'].apply(clean_currency)
        if 'Alíquota' in df.columns:
            df['Alíquota'] = df['Alíquota'].apply(clean_currency)
        if 'CNPJ' in df.columns:
            df['CNPJ'] = df['CNPJ'].astype(str).str.strip()
        if 'Tomador de Serviços' in df.columns:
            df['Tomador de Serviços'] = df['Tomador de Serviços'].astype(str).str.strip()
        if 'Nota' in df.columns:
            df['Nota'] = df['Nota'].astype(str).str.extract(r'(\d+)')[0]

    else:
        # --- Lógica para PDF ---
        col_mapping = {}
        for col in df.columns:
            if 'Nota' in col and 'Valor' not in col:
                col_mapping[col] = 'Nota'
            elif 'Tomador' in col:
                col_mapping[col] = 'Tomador de Serviços'
            elif 'Valor da Nota' in col:
                col_mapping[col] = 'Valor da Nota'
            elif 'ISS' in col.upper() and 'Valor' in col:
                col_mapping[col] = 'Valor ISS'
            elif 'Cr\u00e9dito' in col or 'Crdito' in col:
                col_mapping[col] = 'Valor Crédito'
            elif 'Emiss' in col:
                col_mapping[col] = 'Emissão'
            elif 'Al\u00edq' in col or 'Aliq' in col:
                col_mapping[col] = 'Alíquota'

        df = df.rename(columns=col_mapping)

        if 'Valor da Nota' in df.columns:
            df['Valor da Nota'] = df['Valor da Nota'].apply(clean_currency)

        if 'Valor ISS' in df.columns:
            df['Valor ISS'] = df['Valor ISS'].apply(clean_currency)

        if 'Alíquota' in df.columns:
            df['Alíquota'] = df['Alíquota'].apply(clean_currency)

        if 'Tomador de Serviços' in df.columns:
            df['CNPJ'] = df['Tomador de Serviços'].str.extract(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})')[0]
            df['Tomador de Serviços'] = df['Tomador de Serviços'].str.replace(
                r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', '', regex=True).str.strip()

        if 'Nota' in df.columns:
            df['Nota'] = df['Nota'].astype(str).str.extract(r'(\d+)')[0]

    return df

def format_brl(value):
    """Formata um float como moeda BRL."""
    return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

def generate_text_report(df):
    """Gera um relatório em texto agrupado por Tomador e CNPJ."""
    report = []
    group_cols = ['Tomador de Serviços', 'CNPJ']
    
    # Verifica se as colunas existem
    for col in group_cols:
        if col not in df.columns:
            return "Colunas necessárias não encontradas para gerar o relatório."
            
    # Ordenar para garantir consistência
    df_sorted = df.sort_values(by=group_cols)
    grouped = df_sorted.groupby(group_cols)
    
    for name, group in grouped:
        if isinstance(name, tuple):
            tomador = name[0]
            cnpj = name[1]
        else:
            tomador = name
            cnpj = ""
            
        report.append("=" * 60)
        report.append(f"Tomador de Serviço: {tomador}")
        report.append(f"CNPJ: {cnpj}")
        
        # Totais do grupo
        qtd = len(group)
        valor_total = group['Valor da Nota'].sum() if 'Valor da Nota' in group.columns else 0
        iss_total = group['Valor ISS'].sum() if 'Valor ISS' in group.columns else 0
        
        report.append(f"Qtd. Notas: {qtd}")
        report.append(f"Valor Total: {format_brl(valor_total)}")
        if iss_total > 0:
            report.append(f"ISS Total: {format_brl(iss_total)}")
        
        report.append("Notas fiscais correspondentes:")
        for i, row in enumerate(group.itertuples()):
            nota = getattr(row, 'Nota', 'N/A')
            valor = getattr(row, '_4', 0) if 'Valor da Nota' in group.columns else 'N/A'
            report.append(f"  {i+1}. NFS-e nº {nota}")
        report.append("-" * 60)
        
    return "\n".join(report)

def generate_pdf_report(df):
    """Gera um relatório em PDF agrupado por Tomador e CNPJ."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=10)
    
    group_cols = ['Tomador de Serviços', 'CNPJ']
    
    for col in group_cols:
        if col not in df.columns:
            pdf.cell(200, 10, txt="Colunas necessárias não encontradas para gerar o relatório.", new_x="LMARGIN", new_y="NEXT")
            return pdf.output()
            
    df_sorted = df.sort_values(by=group_cols)
    grouped = df_sorted.groupby(group_cols)
    
    for name, group in grouped:
        if isinstance(name, tuple):
            tomador = name[0]
            cnpj = name[1]
        else:
            tomador = name
            cnpj = ""
            
        # Totais
        qtd = len(group)
        valor_total = group['Valor da Nota'].sum() if 'Valor da Nota' in group.columns else 0
        iss_total = group['Valor ISS'].sum() if 'Valor ISS' in group.columns else 0
        
        pdf.set_font("Helvetica", style="B", size=10)
        pdf.cell(200, 8, txt=f"Tomador: {tomador}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(200, 6, txt=f"CNPJ: {cnpj}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(200, 6, txt=f"Qtd. Notas: {qtd} | Valor Total: {format_brl(valor_total)}", new_x="LMARGIN", new_y="NEXT")
        if iss_total > 0:
            pdf.cell(200, 6, txt=f"ISS Total: {format_brl(iss_total)}", new_x="LMARGIN", new_y="NEXT")
        
        pdf.set_font("Helvetica", style="I", size=9)
        pdf.cell(200, 6, txt="Notas fiscais:", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", size=9)
        
        for i, row in enumerate(group.itertuples()):
            nota = getattr(row, 'Nota', 'N/A')
            pdf.cell(200, 5, txt=f"  {i+1}. NFS-e nr {nota}", new_x="LMARGIN", new_y="NEXT")
            
        pdf.cell(200, 4, txt="-" * 40, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)
        
    return pdf.output()

def generate_excel_report(df):
    """Gera um arquivo Excel formatado com 2 abas: Resumo por CNPJ e Notas Detalhadas."""
    wb = Workbook()

    # =====================================================================
    # ESTILOS COMPARTILHADOS
    # =====================================================================
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    zebra_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    totals_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    totals_font = Font(name='Calibri', bold=True, size=11)
    right_align = Alignment(horizontal='right')
    left_align = Alignment(horizontal='left')

    def apply_header(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # =====================================================================
    # ABA 1: RESUMO POR CNPJ (dados agrupados com somas)
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Resumo por CNPJ"

    headers_resumo = ['Item', 'CNPJ', 'Tomador de Serviços', 'Qtd. Notas',
                       'Valor Total (R$)', 'ISS Total (R$)']
    apply_header(ws1, headers_resumo)

    # Agrupar por CNPJ + Tomador
    group_cols = []
    if 'CNPJ' in df.columns:
        group_cols.append('CNPJ')
    if 'Tomador de Serviços' in df.columns:
        group_cols.append('Tomador de Serviços')

    if group_cols:
        agg_dict = {}
        if 'Nota' in df.columns:
            agg_dict['Qtd_Notas'] = ('Nota', 'count')
        if 'Valor da Nota' in df.columns:
            agg_dict['Valor_Total'] = ('Valor da Nota', 'sum')
        if 'Valor ISS' in df.columns:
            agg_dict['ISS_Total'] = ('Valor ISS', 'sum')

        grouped = df.groupby(group_cols).agg(**agg_dict).reset_index()
        grouped = grouped.sort_values(by=group_cols).reset_index(drop=True)
    else:
        grouped = pd.DataFrame()

    for row_idx, (_, row) in enumerate(grouped.iterrows(), 2):
        # Item
        cell = ws1.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '0'

        # CNPJ
        cnpj_val = str(row.get('CNPJ', '')).strip()
        cell = ws1.cell(row=row_idx, column=2, value=cnpj_val)
        cell.alignment = left_align
        cell.border = thin_border
        cell.number_format = '@'

        # Tomador
        cell = ws1.cell(row=row_idx, column=3, value=str(row.get('Tomador de Serviços', '')))
        cell.alignment = left_align
        cell.border = thin_border

        # Qtd Notas
        qtd = row.get('Qtd_Notas', 0)
        cell = ws1.cell(row=row_idx, column=4, value=int(qtd))
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '0'

        # Valor Total
        val = row.get('Valor_Total', 0)
        try:
            val = float(val)
        except (ValueError, TypeError):
            val = 0.0
        cell = ws1.cell(row=row_idx, column=5, value=val)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = 'R$ #,##0.00'

        # ISS Total
        iss = row.get('ISS_Total', 0)
        try:
            iss = float(iss)
        except (ValueError, TypeError):
            iss = 0.0
        cell = ws1.cell(row=row_idx, column=6, value=iss)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = 'R$ #,##0.00'

        # Zebra
        if row_idx % 2 == 0:
            for c in range(1, len(headers_resumo) + 1):
                ws1.cell(row=row_idx, column=c).fill = zebra_fill

    # Linha de Totais (aba resumo)
    total_row_r = len(grouped) + 2
    cell = ws1.cell(row=total_row_r, column=1, value='TOTAL')
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    for c in range(2, 4):
        cell = ws1.cell(row=total_row_r, column=c)
        cell.fill = totals_fill
        cell.border = thin_border
    # Qtd Notas total
    cell = ws1.cell(row=total_row_r, column=4)
    cell.value = f'=SUBTOTAL(9,D2:D{total_row_r - 1})'
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    cell.number_format = '0'
    cell.alignment = right_align
    # Valor Total
    cell = ws1.cell(row=total_row_r, column=5)
    cell.value = f'=SUBTOTAL(9,E2:E{total_row_r - 1})'
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = right_align
    # ISS Total
    cell = ws1.cell(row=total_row_r, column=6)
    cell.value = f'=SUBTOTAL(9,F2:F{total_row_r - 1})'
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = right_align

    # Tabela + AutoFilter + Freeze (aba resumo)
    if len(grouped) > 0:
        table_ref_r = f'A1:F{total_row_r - 1}'
        tab_r = Table(displayName='tblResumo', ref=table_ref_r)
        tab_r.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2', showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws1.add_table(tab_r)
        ws1.auto_filter.ref = table_ref_r
    ws1.freeze_panes = 'A2'
    for i, w in enumerate([8, 22, 45, 14, 22, 22], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # =====================================================================
    # ABA 2: NOTAS DETALHADAS (cada nota individual)
    # =====================================================================
    ws2 = wb.create_sheet(title="Notas Detalhadas")

    headers_det = ['Item', 'Nota', 'Emissão', 'CNPJ', 'Tomador de Serviços',
                   'Alíquota (%)', 'Valor da Nota (R$)', 'Valor ISS (R$)']
    apply_header(ws2, headers_det)

    # Ordenar
    sort_cols = []
    if 'Tomador de Serviços' in df.columns:
        sort_cols.append('Tomador de Serviços')
    if 'Emissão' in df.columns:
        sort_cols.append('Emissão')
    df_sorted = df.copy()
    if sort_cols:
        ascending = [True] + [False] * (len(sort_cols) - 1)
        df_sorted = df_sorted.sort_values(by=sort_cols, ascending=ascending).reset_index(drop=True)

    for row_idx, (_, row) in enumerate(df_sorted.iterrows(), 2):
        # Item
        cell = ws2.cell(row=row_idx, column=1, value=row_idx - 1)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '0'

        # Nota
        nota_val = row.get('Nota', '')
        try:
            nota_val = int(nota_val)
        except (ValueError, TypeError):
            pass
        cell = ws2.cell(row=row_idx, column=2, value=nota_val)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '0'

        # Emissão
        cell = ws2.cell(row=row_idx, column=3, value=str(row.get('Emissão', '')))
        cell.alignment = left_align
        cell.border = thin_border

        # CNPJ
        cnpj_val = str(row.get('CNPJ', '')).strip()
        cell = ws2.cell(row=row_idx, column=4, value=cnpj_val)
        cell.alignment = left_align
        cell.border = thin_border
        cell.number_format = '@'

        # Tomador
        cell = ws2.cell(row=row_idx, column=5, value=str(row.get('Tomador de Serviços', '')))
        cell.alignment = left_align
        cell.border = thin_border

        # Alíquota
        aliq_val = row.get('Alíquota', 0)
        try:
            aliq_val = float(aliq_val)
        except (ValueError, TypeError):
            aliq_val = 0.0
        cell = ws2.cell(row=row_idx, column=6, value=aliq_val / 100 if aliq_val > 1 else aliq_val)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = '0.00%'

        # Valor da Nota
        valor_val = row.get('Valor da Nota', 0)
        try:
            valor_val = float(valor_val)
        except (ValueError, TypeError):
            valor_val = 0.0
        cell = ws2.cell(row=row_idx, column=7, value=valor_val)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = 'R$ #,##0.00'

        # Valor ISS
        iss_val = row.get('Valor ISS', 0)
        try:
            iss_val = float(iss_val)
        except (ValueError, TypeError):
            iss_val = 0.0
        cell = ws2.cell(row=row_idx, column=8, value=iss_val)
        cell.alignment = right_align
        cell.border = thin_border
        cell.number_format = 'R$ #,##0.00'

        # Zebra
        if row_idx % 2 == 0:
            for c in range(1, len(headers_det) + 1):
                ws2.cell(row=row_idx, column=c).fill = zebra_fill

    # Linha de Totais (aba detalhada)
    total_row_d = len(df_sorted) + 2
    cell = ws2.cell(row=total_row_d, column=1, value='TOTAL')
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    for c in range(2, 7):
        cell = ws2.cell(row=total_row_d, column=c)
        cell.fill = totals_fill
        cell.border = thin_border
    cell = ws2.cell(row=total_row_d, column=7)
    cell.value = f'=SUBTOTAL(9,G2:G{total_row_d - 1})'
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = right_align
    cell = ws2.cell(row=total_row_d, column=8)
    cell.value = f'=SUBTOTAL(9,H2:H{total_row_d - 1})'
    cell.font = totals_font
    cell.fill = totals_fill
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = right_align

    # Tabela + AutoFilter + Freeze (aba detalhada)
    if len(df_sorted) > 0:
        table_ref_d = f'A1:H{total_row_d - 1}'
        tab_d = Table(displayName='tblNotasFiscais', ref=table_ref_d)
        tab_d.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium2', showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws2.add_table(tab_d)
        ws2.auto_filter.ref = table_ref_d
    ws2.freeze_panes = 'A2'
    for i, w in enumerate([8, 12, 22, 22, 45, 14, 22, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- Salvar em buffer ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# === INTERFACE ===
st.title("📊 Analisador de Notas Fiscais")
st.caption("Versão 2.1.0 — Upload múltiplo")
st.markdown("Faça o upload de até **12 arquivos** (PDF ou CSV) para extrair, analisar e agrupar os dados por **Tomador de Serviços** e **CNPJ**.")

# --- Estado da sessão ---
if 'uploader_key' not in st.session_state:
    st.session_state.uploader_key = 0

# --- Upload múltiplo ---
uploaded_files = st.file_uploader(
    "Arraste e solte seus arquivos PDF ou CSV aqui (máx. 12)",
    type=["pdf", "csv"],
    accept_multiple_files=True,
    key=f"uploader_{st.session_state.uploader_key}"
)

# Limitar a 12 arquivos
if uploaded_files and len(uploaded_files) > 12:
    st.warning("⚠️ Máximo de 12 arquivos permitidos. Apenas os 12 primeiros serão processados.")
    uploaded_files = uploaded_files[:12]

if uploaded_files:
    # Mostrar arquivos carregados em grid
    st.subheader(f"📎 {len(uploaded_files)} arquivo(s) carregado(s)")
    
    # Grid de 4 colunas para mostrar os arquivos
    cols_per_row = 4
    for i in range(0, len(uploaded_files), cols_per_row):
        cols = st.columns(cols_per_row)
        for j, col in enumerate(cols):
            idx = i + j
            if idx < len(uploaded_files):
                f = uploaded_files[idx]
                icon = "📄" if f.name.lower().endswith('.pdf') else "📊"
                size_kb = len(f.getvalue()) / 1024
                col.markdown(
                    f"""<div style="background:#1a1a2e;border:1px solid #333;border-radius:8px;padding:12px;text-align:center;">
                    <span style="font-size:28px;">{icon}</span><br>
                    <strong style="color:#e0e0e0;font-size:13px;">{f.name[:25]}</strong><br>
                    <span style="color:#888;font-size:11px;">{size_kb:.0f} KB</span>
                    </div>""",
                    unsafe_allow_html=True
                )

    st.divider()

    # Botões de ação
    col_analisar, col_nova = st.columns([3, 1])
    analisar = col_analisar.button("🔍 Analisar Todos os Arquivos", use_container_width=True, type="primary")
    
    if col_nova.button("🔄 Nova Análise", use_container_width=True):
        st.session_state.uploader_key += 1
        st.rerun()

    if analisar:
        try:
            all_dfs = []
            progress = st.progress(0, text="Processando arquivos...")

            for idx, f in enumerate(uploaded_files):
                progress.progress((idx + 1) / len(uploaded_files),
                                  text=f"Processando {f.name} ({idx+1}/{len(uploaded_files)})...")

                if f.type == "application/pdf" or f.name.lower().endswith('.pdf'):
                    raw_df = process_pdf(f.getvalue())
                else:
                    raw_df = load_csv(f)

                if not raw_df.empty:
                    df_clean = prepare_dataframe(raw_df)
                    df_clean['_Arquivo'] = f.name  # rastreabilidade
                    all_dfs.append(df_clean)
                else:
                    st.warning(f"⚠️ Nenhum dado encontrado em: {f.name}")

            progress.empty()

            if not all_dfs:
                st.error("Nenhum dado válido encontrado nos arquivos enviados.")
            else:
                # Combinar todos os DataFrames
                df = pd.concat(all_dfs, ignore_index=True)

                st.success(f"✅ Extração concluída! **{len(df)} notas** extraídas de **{len(all_dfs)} arquivo(s)**.")

                st.divider()

                # --- KPIs ---
                kpi_cols = st.columns(5)
                with kpi_cols[0]:
                    st.metric("Arquivos", len(all_dfs))
                with kpi_cols[1]:
                    st.metric("Total de Notas", len(df))
                with kpi_cols[2]:
                    if 'Valor da Nota' in df.columns:
                        st.metric("Valor Total Faturado", format_brl(df['Valor da Nota'].sum()))
                with kpi_cols[3]:
                    if 'Tomador de Serviços' in df.columns:
                        st.metric("Tomadores Únicos", df['Tomador de Serviços'].nunique())
                with kpi_cols[4]:
                    if 'Valor ISS' in df.columns:
                        st.metric("Total ISS", format_brl(df['Valor ISS'].sum()))

                st.divider()

                # --- VISÃO AGRUPADA ---
                st.subheader("📁 Agrupamento por Tomador (CNPJ)")

                df_filtered = df.copy()
                tomador_selecionado = "Todos"

                if 'Tomador de Serviços' in df.columns and 'CNPJ' in df.columns:
                    tomadores = sorted(df['Tomador de Serviços'].dropna().unique())

                    tomador_selecionado = st.selectbox(
                        "Selecione um Tomador de Serviços para detalhar:",
                        ["Todos"] + list(tomadores)
                    )

                    df_filtered = df if tomador_selecionado == "Todos" else df[df['Tomador de Serviços'] == tomador_selecionado]

                    # Agrupamento por Tomador + CNPJ
                    group_cols = ['Tomador de Serviços', 'CNPJ']

                    grouped = df_filtered.groupby(group_cols).agg(**{
                        'Qtd_Notas': ('Nota', 'count'),
                        **({'Valor_Total': ('Valor da Nota', 'sum')} if 'Valor da Nota' in df.columns else {}),
                        **({'ISS_Total': ('Valor ISS', 'sum')} if 'Valor ISS' in df.columns else {}),
                    }).reset_index()

                    grouped = grouped.sort_values(by=group_cols)
                    df_filtered = df_filtered.sort_values(by=group_cols)

                    # Formatar moeda para exibição
                    display_grouped = grouped.copy()
                    if 'Valor_Total' in display_grouped.columns:
                        display_grouped['Valor Total'] = display_grouped['Valor_Total'].apply(format_brl)
                        display_grouped = display_grouped.drop(columns=['Valor_Total'])
                    if 'ISS_Total' in display_grouped.columns:
                        display_grouped['ISS Total'] = display_grouped['ISS_Total'].apply(format_brl)
                        display_grouped = display_grouped.drop(columns=['ISS_Total'])

                    st.dataframe(
                        display_grouped,
                        use_container_width=True,
                        hide_index=True
                    )

                    # Detalhe das Notas
                    with st.expander("Ver lista detalhada de notas do filtro atual"):
                        st.dataframe(df_filtered, use_container_width=True)

                else:
                    st.warning("Não foi possível identificar as colunas de 'Tomador de Serviços' ou 'CNPJ' na tabela extraída.")
                    st.write("Dados brutos extraídos:")
                    st.dataframe(df)

                # --- EXPORTAÇÃO ---
                st.divider()
                st.subheader("📥 Exportar Dados")

                col_txt, col_pdf, col_xlsx = st.columns([1, 1, 1])

                nome_arquivo_tomador = "Todos" if tomador_selecionado == "Todos" else tomador_selecionado.replace(" ", "_").replace("/", "").replace(".", "")

                # Remover coluna _Arquivo antes de exportar
                df_export = df_filtered.drop(columns=['_Arquivo'], errors='ignore')

                report_txt = generate_text_report(df_export)
                col_txt.download_button(
                    label="Baixar Relatório (TXT)",
                    data=report_txt,
                    file_name=f'analise_{nome_arquivo_tomador[:20]}.txt',
                    mime='text/plain',
                    use_container_width=True
                )

                report_pdf = generate_pdf_report(df_export)
                col_pdf.download_button(
                    label="Baixar Relatório (PDF)",
                    data=bytes(report_pdf),
                    file_name=f'analise_{nome_arquivo_tomador[:20]}.pdf',
                    mime='application/pdf',
                    use_container_width=True
                )

                report_xlsx = generate_excel_report(df_export)
                col_xlsx.download_button(
                    label="📊 Baixar Planilha (Excel)",
                    data=report_xlsx,
                    file_name=f'notas_fiscais_{nome_arquivo_tomador[:20]}.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"Ocorreu um erro durante o processamento: {str(e)}")

